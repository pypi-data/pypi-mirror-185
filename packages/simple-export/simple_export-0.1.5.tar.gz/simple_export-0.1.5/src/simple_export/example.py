#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2022/12/10 0:33
# @Author  : mtl
# @File    : example.py
# @Description : *****
import traceback

from openpyxl import load_workbook

from simple_export.excel import write_excel_for_template

def test1():
    wb_tmp = load_workbook(f'./template/excel1.xlsx')
    value = {
        "家庭财产库存清单": {
            "thing": [
                {
                    "id": "=ROW($A1)",
                    "room": "客厅",
                    "goods": "物品1",
                    "structure": "制造商1",
                    "serial_number": "33XCBH3",
                    "time": "=TODAY()-120",
                    "source": "联机",
                    "price": "2000",
                    "eval_price": "2000",
                    "remark": "",
                    "has_photo": "是"
                }
            ],
            "test": "1"
        },
        "sheet1": {
            "data": {
                "name": "1",
                "data_list": [
                    {"a": 3},
                    {"a": 4}
                ],
                "b": "cec测试"
            },
            "data_list": [
                {
                    "a": 1,
                    "b": 2
                },
                {
                    "a": 2,
                    "b": 1
                }
            ],
            "name": 123
        },
        "sheet2": {
            "data": {
                "name": "1",
                "data_list": [
                    {"a": 3}
                ],
                "b": "cec测试"
            }
        }
    }
    write_excel_for_template(value=value, wb_tmp=wb_tmp)
    wb_tmp.save("./val1.xlsx")
    wb_tmp.close()
    wb_tmp = load_workbook(f'./val1.xlsx', data_only=True)


def test2():
    wb_tmp = load_workbook(f'./template/excel2.xlsx')
    value = {
        "sheet1": {
            "data": {
                "name": "1",
                "data_list": [
                    {"a": 3},
                    {"a": 4}
                ],
                "b": "cec测试"
            },
            "data_list": [
                {
                    "a": 1,
                    "b": 2
                },
                {
                    "a": 2,
                    "b": 1
                }
            ],
            "name": 123
        },
        "Sheet2": {
            "data": {
                "b": "1"
            },
            "b": "cec测试"
        }
    }
    write_excel_for_template(value=value, wb_tmp=wb_tmp)
    wb_tmp.save("./val2.xlsx")
    # wb_tmp.close()
    wb_tmp = load_workbook(f'./vaL2.xlsx', data_only=True, read_only=True)

if __name__ == '__main__':
    test1()
    test2()