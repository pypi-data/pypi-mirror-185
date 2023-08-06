from .observer import Observer
from ...events.export_events import ExcelExportEvent
from ...tools import DataFrameTools as DfTools
import openpyxl.utils as pyx_tools
from typing import Any


# output of the progress for running the report
PROGRESS_OUTPUT_FILE = "progress.txt"


# ExcelExportObserver: Observer to format exported excel files
class ExcelExportObserver(Observer):
    def __init__(self):
        super().__init__()
        self.excel = None


    # notify_export(out_event): Formats the exported excel files
    def notify_export(self, out_event: ExcelExportEvent):
        # write all our dataframes into the excel sheet
        export_info = out_event.export()
        self.excel = export_info.excel_tool
        writer = export_info.writer
        sheet = export_info.sheet

        # freeze the top row
        self.excel.freeze_rows(sheet, 1)

        # autofit the column widths
        width_lst = DfTools.col_get_max_width(out_event.df_lst[0].df, with_header = True)
        self.excel.change_col_widths(sheet, width_lst, 2, sheet.max_column)

        # apply filter to the columns
        filter_area = f"A1:{pyx_tools.get_column_letter(sheet.max_column)}{sheet.max_row}"
        sheet.auto_filter.ref = filter_area

        self.excel.save_current_file()


    # notify(out_event): Formats the exported excel files based off, 'target'
    def notify(self, target: Any):
        if (isinstance(target, ExcelExportEvent)):
            self.notify_export(target)
