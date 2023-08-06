import openpyxl as pyx
import enum
import openpyxl.utils as pyx_tools
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, Alignment, Border, PatternFill
from .colours import ColourList
from typing import List, Any, Tuple, Union, Optional, Callable, Tuple


RANGE_SEPERATOR = ":"

# ExcelAlignment: Basic alignments for an excel
class ExcelAlignment(enum.Enum):
    Left = "left"
    Right = "right"
    Center = "center"
    Justify = "justify"


# ExcelStyles: Class to deal with all excel styles
class ExcelStyles():
    def __init__(self, alignment: Optional[Alignment] = None, font: Optional[Font] = None, number_format: Optional[str] = None,
                 border: Optional[Border] = None, fill: Optional[PatternFill] = None, is_merged_cell: bool = False):
        self.alignment = alignment
        self.font = font
        self.border = border
        self.fill = fill
        self.number_format = number_format
        self.is_merged_cell = is_merged_cell


    # format_merged(sheet, cell_range): Formats the borders for a merged cell range
    def format_merged(self, table: Tuple[Tuple[Cell]]):
        top_row = table[0]
        bottom_row = table[-1]
        border_format_exist = bool(self.border is not None)

        if (border_format_exist):
            # retrieve border formatting for all 4 sides
            left_border = Border(left = self.border.left)
            right_border = Border(right = self.border.right)
            top_border = Border(top = self.border.top)
            bottom_border = Border(bottom = self.border.bottom)

        # add the side borders and add the fill to the cells
        for row in table:
            left = row[0]
            right = row[-1]

            # add the side borders
            if (border_format_exist):
                left.border = left.border + left_border
                right.border = right.border + right_border

            # add the fill
            if (self.fill is not None):
                for cell in row:
                    cell.fill = self.fill

        # add the top/bottom borders
        if (border_format_exist):
            for cell in top_row:
                cell.border = cell.border + top_border

            for cell in bottom_row:
                cell.border = cell.border + bottom_border


    # cell_apply(cell): Applies the different styles to 'cell'
    def cell_apply(self, area: Union[Cell, Tuple[Tuple[Cell]]]):
        # for merged cells, get corresponding cell on the top left corner of the
        #   merged selection and retrieve any other necessary parameters
        if (self.is_merged_cell):
            cell = area[0][0]
        else:
            cell = area

        # format the cell
        if (self.alignment is not None):
            cell.alignment = self.alignment

        if (self.font is not None):
            cell.font = self.font

        if (self.number_format is not None):
            cell.number_format = self.number_format

        if (self.is_merged_cell):
            self.format_merged(area)
        else:
            if (self.fill is not None):
                cell.fill = self.fill

            if (self.border is not None):
                cell.border = self.border


    # group_apply(table): Applies styles to a group of cells
    def group_apply(self, table: Tuple[Tuple[Cell]]):
        for row in table:
            for cell in row:
                self.apply(cell)


    # apply(area): Applies the necessary styles to 'area'
    def apply(self, area: Union[Cell, Tuple[Tuple[Cell]]]):
        if (isinstance(area, Tuple) and not self.is_merged_cell):
            self.group_apply(area)
        else:
            self.cell_apply(area)


# ExcelTableStyles: Different excel table formatting styles
class ExcelTableStyles(enum.Enum):
    MediumBlack1 = "TableStyleMedium1"
    MediumBlack2 = "TableStyleMedium8"
    MediumBlack3 = "TableStyleMedium15"
    MediumBlack4 = "TableStyleMedium22"
    MediumBlue1 = "TableStyleMedium2"
    MediumBlue2 = "TableStyleMedium9"
    MediumBlue3 = "TableStyleMedium16"
    MediumBlue4 = "TableStyleMedium23"


# Excel: Class to handle reading/writing of excel files
class Excel():
    MIN_ROW_VAL = 1
    DEFAULT_FIRST_SHEET_NAME = "Sheet1"

    def __init__(self, loc: str):
        self.loc = loc
        self.workbook = None
        self.newly_created = False


    # open_file() Opens a excel file
    def open_file(self) -> Workbook:
        wb = pyx.load_workbook(filename=self.loc)
        self.workbook = wb
        return wb


    # read_sheet(sheet_name) Gets the sheet by the name 'sheet_name'
    #   from the designated excel file
    # requires: self.workbook is not None
    # effects: may end the program
    def read_sheet(self, sheet_name: str) -> Worksheet:
        if (self.workbook is None):
            self.open_file()

        sheet = self.workbook[sheet_name]
        return sheet


    # create_sheet(sheet_name): Creates a new worksheet by the name 'sheet_name'
    def create_sheet(self, sheet_name: str) -> Worksheet:
        worksheet = self.workbook.create_sheet(sheet_name)
        return worksheet


    # sheet_action(func): Decorator to open up a sheet before calling 'func'
    def sheet_action(func: Callable[..., Any]):
        def check_sheet(self, sheet: Union[str, Worksheet], *args, **kwargs) -> Any:
            if (isinstance(sheet, str)):
                sheet = self.read_sheet(sheet)

            return func(self, sheet, *args, **kwargs)
        return check_sheet


    # read_row(sheet_no, row_index) Reads in a row from an excel sheet
    # note: if 'sheet' has data type Worksheet,
    #       then 'sheet' does not necessarily need to be a sheet from the current
    #       excel file
    @sheet_action
    def read_row(self, sheet: Union[str, Worksheet],
                 row_index: int) -> Tuple[pyx.cell.cell.Cell]:
        row = sheet[row_index]
        return row


    # save_file(wb) Saves an excel file
    # note: 'wb' does not neccessarily need to be the current opened workbook,
    #   self.workbook
    def save_file(self, wb: Workbook):
        wb.save(self.loc)
        return wb


    # save_current_file() Saves the current opened excel file
    # requires: self.workbook is not None
    def save_current_file(self):
        return self.save_file(self.workbook)


    # create_file() Cretes a new excel file based off the given location
    def create_file(self, first_sheet_name: Optional[str] = None) -> Workbook:
        wb = pyx.Workbook()
        wb.save(self.loc)
        self.newly_created = True

        if (first_sheet_name is not None):
            wb[self.DEFAULT_FIRST_SHEET_NAME].title = sheet_name

        self.workbook = wb
        return wb


    # create_sheet(name) Creates a new sheet from the current excel file
    # requires: self.workbook is not None
    # effects: user must save the excel file
    def create_sheet(self, name: str) -> Worksheet:
        self.workbook.create_sheet(name)
        return self.workbook[name]


    # write_row(sheet, row_index, row_values) Write the values from 'row_values'
    #   into some row in 'sheet' of some excel file
    # note: if 'sheet' has data type Worksheet,
    #       then 'sheet' does not necessarily need to be a sheet from the current
    #       excel file
    # effects: may end the program
    @sheet_action
    def write_row(self, sheet: Union[str, Worksheet],
                  row_index: int, row_values: List[str], font: Optional[pyx.styles.Font] = None):

        row_values_len = len(row_values)
        for i in range(row_values_len):
            cell = sheet.cell(row=row_index, column=(i + 1))
            if (font is not None):
                cell.font = font
            cell.value = row_values[i]

        if (self.newly_created):
            self.newly_created = False


    # add_row(sheet, row_values) Write the values from 'row_values'
    #   into the next new row in 'sheet' of some excel file
    # note: if 'sheet' has data type Worksheet,
    #       then 'sheet' does not necessarily need to be a sheet from the current
    #       excel file
    # effects: may end the program
    @sheet_action
    def add_row(self, sheet: Union[str, Worksheet],
                row_values: List[str], font: Optional[pyx.styles.Font] = None):

        row_index = sheet.max_row
        if (not self.newly_created):
            row_index += 1
        else:
            self.newly_created = False

        self.write_row(sheet, row_index, row_values)


    # write_headings(sheet, heading_lst) Writes the heading to the first row
    #   in 'sheet' of some excel file
    # note: if 'sheet' has data type Worksheet,
    #       then 'sheet' does not necessarily need to be a sheet from the current
    #       excel file
    @sheet_action
    def write_headings(self, sheet: Union[str, Worksheet],
                       heading_lst: List[str]):

        font = pyx.styles.Font(bold=True)
        self.write_row(sheet, self.MIN_ROW_VAL, heading_lst, font = font)


    # remove_columns(sheet, start_ind, no_of_columns): Removes columns from 'sheet'
    # requires: start_ind > 0
    #           no_of_columns > 0
    @sheet_action
    def remove_columns(self, sheet: Union[str, Worksheet], start_ind: int, no_of_columns: int = 1):
        sheet.delete_cols(start_ind, no_of_columns)


    # insert_rows(sheet, start_ind, no_of_columns): inserts rows from 'sheet'
    # requires: start_ind > 0
    #           no_of_rows > 0
    @sheet_action
    def insert_rows(self, sheet: Union[str, Worksheet], start_ind: int, no_of_rows: int = 1):
        sheet.insert_rows(start_ind, no_of_rows)



    # insert_columns(sheet, start_ind, no_of_columns): inserts columns from 'sheet'
    # requires: start_ind > 0
    #           no_of_rows > 0
    @sheet_action
    def insert_columns(self, sheet: Union[str, Worksheet], start_ind: int, no_of_columns: int = 1):
        sheet.insert_cols(start_ind, no_of_columns)


    # format_table_headings(sheet, row, left_col, right_col): Formats the headings
    #   of a table
    # requires: row >= 1
    #           right_col >= left_col >=  1
    @sheet_action
    def format_table_headings(self, sheet: Union[str, Worksheet], row: int, left_col: int, right_col: int, styles: ExcelStyles):
        right_col += 1

        for i in range(left_col, right_col):
            cell = sheet.cell(row = row, column = i)
            styles.apply(cell)


    # get_range(left_top, right_bottom): Formats a range value on excel
    #  requires: format for 'left_top' and 'right_bottom' is [Capital Letter][number]
    #                ex. A8
    @classmethod
    def get_range(cls, left_top: str, right_bottom: str) -> str:
        return f"{left_top}:{right_bottom}"


    # get_row_range(left, right, row): Retrieves the cell range for a certain row
    @classmethod
    def get_row_range(cls, left: str, right: str, row: int) -> str:
        if (left == right):
            return f"{left}{row}"
        else:
            return cls.get_range(f"{left}{row}", f"{right}{row}")


    # get_col_range(top, bottom, col): Retrieves the cell range for a certain column
    # Requires: top <= bottom
    @classmethod
    def get_col_range(cls, top: int, bottom: int, col: str):
        if (top == bottom):
            return f"{col}{top}"
        else:
            return cls.get_range(f"{col}{top}", f"{col}{bottom}")


    # get_row_cell_lst(col_lst, row_no): Retrieves the string representation of
    #    list of cells forming a row
    @classmethod
    def get_row_cell_lst(cls, col_lst: List[str], row_no: int) -> str:
        col_lst_len = len(col_lst)
        result = ""

        for i in range(col_lst_len):
            if (i):
                result += ","

            result += f"{col_lst[i]}{row_no}"

        return result


    # get_cell_lst(cell_lst): Retrieves the string representation of
    #   a list of cells
    @classmethod
    def get_cell_lst(cls, cell_lst: List[Tuple[Union[str, int]]]) -> str:
        cell_lst_len = len(cell_lst)
        result = ""

        for i in range(cell_lst_len):
            if (i):
                result += ","

            result += f"{cell_lst[i][0]}{cell_lst[i][1]}"

        return result


    # get_col_cell_lst(row_lst, col): Retrieves the string representation of the
    #   list of cells forming a column
    @classmethod
    def get_col_cell_lst(cls, row_lst: List[Union[str, int]], col: str) -> str:
        row_lst_len = len(row_lst)
        result = ""

        for i in range(row_lst_len):
            if (i):
                result += ","
            result += f"{col}{row_lst[i]}"

        return result


    # create_table(sheet, table_name, left_top, right_bottom): Creates a table on 'sheet'
    #  requires: format for 'left_top' and 'right_bottom' is [Capital Letter][number]
    #                ex. A8
    @sheet_action
    def create_table(self, sheet: Union[str, Worksheet], table_name: str, left_top: str, right_bottom: str, header_styles: Optional[ExcelStyles] = None):
        table = Table(displayName = table_name, ref = self.get_range(left_top, right_bottom))

        # basic style for the table
        style = TableStyleInfo(name=ExcelTableStyles.MediumBlue1.value, showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        sheet.add_table(table)

        # format the headings for the table
        if (header_styles is not None):
            left_top = pyx_tools.cell.coordinate_from_string(left_top)
            right_bottom = pyx_tools.cell.coordinate_from_string(right_bottom)

            left_col = pyx_tools.column_index_from_string(left_top[0])
            right_col = pyx_tools.column_index_from_string(right_bottom[0])
            self.format_table_headings(sheet, left_top[1], left_col, right_col, header_styles)


    # change_col_widths(sheet, width_lst, left, right): Changes the column widths
    #    of 'sheet' based from 'width_lst'
    # requires: sheet.max_column >= right >= left >= 1
    #           len(width_lst) == right - left
    @sheet_action
    def change_col_widths(self, sheet: Union[str, Worksheet], width_lst: List[float], left: int, right: int):
        for i, width in enumerate(width_lst, start = left):
            sheet.column_dimensions[pyx_tools.get_column_letter(i)].width = width


    # merge_cells(sheet, left_top, right_bottom): Merges the cells from the range
    #   'left_top' to 'right_bottom'
    #  requires: format for 'left_top' and 'right_bottom' is [Capital Letter][number]
    #                ex. A8
    # Note: To reference the merged cell, refer to the top left cell of the merged
    #   range
    @sheet_action
    def merge_cells(self, sheet: Union[str, Worksheet], left_top:str, right_bottom: str,
                    font: Optional[Font] = None, alignment: Optional[Alignment] = None, border: Optional[Border] = None, fill: Optional[PatternFill] = None):
        # merge the cells
        range = self.get_range(left_top, right_bottom)
        sheet.merge_cells(range)

        # apply styles
        styles = ExcelStyles(alignment = alignment, font = font, border = border, fill = fill, is_merged_cell = True)
        styles.apply(sheet[range])


    # freeze_rows(sheet, no_of_rows): Freezes the top few rows on a sheet
    @sheet_action
    def freeze_rows(self, sheet: Union[str, Worksheet], no_of_rows: int):
        sheet.freeze_panes = f"A{no_of_rows + 1}"


    # freeze_panes(sheet, no_of_rows, no_of_columns): Freezes a group of rows and columns
    #   on a sheet
    @sheet_action
    def freeze_panes(self, sheet: Union[str, Worksheet], no_of_rows: int, no_of_columns: int):
        # get the top left corner of the unfrozen cells
        column = pyx_tools.get_column_letter(no_of_columns + 1)
        top_left_corner = f"{column}{no_of_rows + 1}"

        # freeze the cells
        sheet.freeze_panes = top_left_corner


    # hide_column(sheet, col): Hides the column 'col' in 'sheet'
    @sheet_action
    def hide_column(self, sheet: Union[str, Worksheet], col: str):
        sheet.column_dimensions[col].hidden = True


    # hide_row(sheet, row): Hides the row, 'row', in 'sheet'
    @sheet_action
    def hide_row(self, sheet: Union[str, Worksheet], row: int):
        sheet.row_dimensions[row].hidden = True
