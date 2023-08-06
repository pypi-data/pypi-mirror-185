import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.formatting.rule import CellIsRule


def create_wb():
    wb = openpyxl.Workbook()
    return wb


def create_sheet(wb, sheet_name, index):
    wb.create_sheet(title=sheet_name, index=index)
    target_sheet = wb[sheet_name]
    return target_sheet


def just_open(filename):
    import win32com.client
    xl_app = win32com.client.Dispatch('Excel.Application')
    xl_app.Visible = False
    xl_book = xl_app.Workbooks.Open(filename)
    xl_book.Save()
    xl_book.Close()


def ExcelSizeReduction(file_path, file_name, SheetName, new_file_name):
    import win32com.client
    excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbook.Open(r'{}\{}'.format(file_path, file_name))
    excel.DisplayAlerts = False
    wb.DoNotPromptForConvert = False
    wb.CheckCompatibility = False
    wb.SaveAs(r'{}\{}.xlsx'.format(file_path, new_file_name), FileFormat=51, ConflictResolution=2)
    excel.Application.Quit()
    df = pd.read_excel(r'{}\{}.xlsx'.format(file_path, new_file_name), sheet_name=SheetName)
    df.to_excel(r'{}\{}.xlsx'.format(file_path, new_file_name), index=True)
    wb_new = openpyxl.load_workbook(r'{}\{}.xlsx'.format(file_path, new_file_name))
    wb_new_sh = wb_new['Sheet1']
    return wb_new_sh


class PdfAutomation:
    def __init__(self, target_sheet, heading1_color='749CBA', heading2_color='E0E3E4', text_size_h=12,
                 comment_color='DBECF0'):
        self.target_sheet = target_sheet
        self.heading1_color = heading1_color
        self.heading2_color = heading2_color
        self.text_size_h = text_size_h
        self.font = Font(name='Arial', size=self.text_size_h, bold=True, italic=False, vertAlign=None,
                         color='00FFFFFF')
        self.thin = Side(border_style='thin', color='000000')
        self.vertical_text = 'Side Vertical Text'
        self.vertical_text_cover_page = 'Cover Page Side Vertical Text'
        self.comment_color = comment_color
        """
        :param target_sheet: this is a variable that refers to the excel file on which we want to perform operation.
        :param heading1_color: this is a variable of color code for primary heading of the table
        :param heading2_color: this is a variable of color code for secondary heading of the table        
        """

    def adjust_columns(self, col_width):
        for i in col_width:
            self.target_sheet.column_dimensions[i].width = col_width[i]
        """
        With the help of this function we can adjust the columns width of a individual excel sheet.
        """

    def data_copy_paste(self, ref_sheet):
        for i in range(1, ref_sheet.max_row + 1):
            for j in range(1, ref_sheet.max_column + 1):
                self.target_sheet.cell(row=i, column=j).value = ref_sheet.cell(row=i, column=j).value
        """
        this is for copy pasting the data of reference excel sheet to the target sheet.
        """

    def color_filling(self, s_row, s_col, n_col, n_row, colored_row2, skip_row):
        for i in range(n_col):
            self.target_sheet.cell(row=s_row, column=s_col + i).fill = PatternFill('solid',
                                                                                   start_color=self.heading1_color)
            self.target_sheet.cell(row=s_row, column=s_col + i).font = self.font
            if n_row >= 1:
                if colored_row2 > 0:
                    for j in range(skip_row, colored_row2):
                        self.target_sheet.cell(row=s_row + 1 + j, column=s_col + i).alignment = Alignment(
                            horizontal='center', vertical='center', wrap_text=True)
                        self.target_sheet.cell(row=s_row + 1 + j, column=s_col + i).fill = PatternFill('solid',
                                                                                                       start_color=self.heading2_color)
                    for j in range(colored_row2):
                        self.target_sheet.cell(row=s_row + 1 + j, column=s_col + i).font = Font(name='Arial', size=11,
                                                                                                bold=True)
        """
        With the help of this function we can fill the color at the respective cell or the range of the cell at 
        excel file.
        This function will helping us to create a table via color filling and applying some formatting on it.
        """

    def additional_color_filling(self, s_row, s_col, n_col, color_code, n_row=False, wrap_text=False):
        if not n_row:
            for i in range(n_col):
                j, k = 0, 0
                if n_row:
                    j = i
                else:
                    k = i
                self.target_sheet.cell(row=s_row + j, column=s_col + k).fill = PatternFill('solid',
                                                                                           start_color=color_code)
                self.target_sheet.cell(row=s_row + j, column=s_col + k).font = Font(name='Arial', size=11, bold=True)
                if wrap_text:
                    self.target_sheet.cell(row=s_row + j, column=s_col + k).alignment = Alignment(horizontal='center',
                                                                                                  vertical='center',
                                                                                                  wrap_text=True)
        """
        by the help of this function we can fill the any color in the fixed specific cells either in a vertical
        or a horizontal direction.
        """

    def comments_formate(self, cell_loc):
        for i in cell_loc:
            self.target_sheet[i].font = Font(name='Arial', size=8, italic=True)

    def side_vertical_text(self, s_row, s_col, e_row, cover_page=False):
        self.target_sheet.merge_cells(start_row=s_row, end_row=e_row, s_column=s_col, end_column=s_col)
        if not cover_page:
            self.target_sheet.cell(row=s_row, column=s_col).value = self.vertical_text
        else:
            self.target_sheet.cell(row=s_row, column=s_col).value = self.vertical_text_cover_page
        self.target_sheet.cell(row=s_row, column=s_col).alignment = Alignment(text_rotation=90)
        for i in range(s_row, e_row + 1):
            self.target_sheet.cell(row=i, column=s_col).border = Border(right=self.thin)

    def merging_cells(self, s_row, s_col, e_row, e_col, n_row, columns=False):
        for i in range(n_row):
            j, k = 0, 0
            if not columns:
                j = i
            else:
                k = i
            self.target_sheet.merge_cells(start_row=s_row + j, end_row=e_row + j, s_column=s_col + k,
                                          end_column=e_col + k)

    def data_range_copy_paste(self, ref_sheet, s_row, s_col, n_row, n_col, r_diff, c_diff, first_col=False,
                              wrap_text=False, numeric_col_alignment_disable=0):
        for i in range(n_row):
            for j in range(n_col):
                self.target_sheet.cell(row=s_row + i, column=s_col + j).value = ref_sheet.cell(row=s_row + i - r_diff,
                                                                                               column=s_col + j - c_diff).value
                self.target_sheet.cell(row=s_row + i, column=s_col + j).number_format = '#,##0_);(#,##0);-'
                if j > numeric_col_alignment_disable:
                    self.target_sheet.cell(row=s_row + i, column=s_col + j).alignment = Alignment(horizontal='center',
                                                                                                  vertical='center')
                if first_col and wrap_text:
                    self.target_sheet.cell(row=s_row + i, column=s_col + j).alignment = Alignment(horizontal='center',
                                                                                                  vertical='center',
                                                                                                  wrap_text=True)

    def border_creation(self, s_row, s_col, n_col):
        for i in range(n_col):
            self.target_sheet.cell(row=s_row, column=s_col + i).border = Border(bottom=self.thin)

    def date_formatting(self, s_row, s_col, n_row, format_):
        for i in range(n_row):
            self.target_sheet.cell(row=s_row + i, column=s_col).number_format = format_
            self.target_sheet.cell(row=s_row + i, column=s_col).alignment = Alignment(horizontal='center',
                                                                                      vertical='center')
            self.target_sheet.cell(row=s_row + i, column=s_col).font = Font(name='Arial', size=11, bold=True)

    def comments_entry(self, s_row, s_col, n_row, n_col, merge=True):
        self.target_sheet.cell(row=s_row, column=s_col).value = 'Comment'
        for i in range(n_row):
            for j in range(n_col):
                self.target_sheet.cell(row=s_row + 1 + i, column=s_col + j).fill = PatternFill('Solid',
                                                                                               start_color=self.comment_color)
                self.target_sheet.cell(row=s_row + i, column=s_col + j).fill = PatternFill('Solid',
                                                                                           start_color=self.comment_color)
        self.target_sheet.cell(row=s_row, column=s_col).font = Font(name='Arial', size=11, bold=True,
                                                                    underline='Single')
        self.target_sheet.cell(row=s_row + 1, column=s_col).alignment = Alignment(vertical='top', wrap_text=True)
        if merge:
            self.target_sheet.merge_cells(start_row=s_row + 1, end_row=s_row + n_row + 1, start_col=s_col,
                                          end_column=s_col + n_col - 1)
            self.target_sheet.merge_cells(start_row=s_row, end_row=s_row, start_col=s_col,
                                          end_column=s_col + n_col - 1)

    def conditional_formatting(self, cell_range, min_g, max_g, min_r, max_r, min_o, max_o):
        green_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        red_fill = PatternFill(start_color='EE1111', end_color='EE1111', fill_type='solid')
        orange_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        self.target_sheet.conditional_formatting.add(cell_range, CellIsRule(operator='between', formula=[min_g, max_g],
                                                                            fill=green_fill))
        self.target_sheet.conditional_formatting.add(cell_range, CellIsRule(operator='between', formula=[min_o, max_o],
                                                                            fill=orange_fill))
        self.target_sheet.conditional_formatting.add(cell_range, CellIsRule(operator='between', formula=[min_r, max_r],
                                                                            fill=red_fill))

    def number_formatting(self, s_row, s_col, n_row, n_col, formate, h_alignment, v_alignment):
        for i in range(n_row):
            for j in range(n_col):
                self.target_sheet.cell(row=s_row + i, column=s_col + j).alignment = Alignment(horizontal=h_alignment,
                                                                                              vertical=v_alignment)
                if formate != 0:
                    self.target_sheet.cell(row=s_row + i, column=s_col + j).number_format = formate

    def box_border_creation(self, s_row, e_row, s_col, e_col):
        for row in self.target_sheet['{}{}:{}{}'.format(s_col, e_row, e_col, e_row)]:
            for cell in row:
                cell.border = Border(bottom=self.thin)
        for row in self.target_sheet['{}{}:{}{}'.format(e_col, s_row, e_col, e_row)]:
            for cell in row:
                cell.border = Border(bottom=self.thin)
        for row in self.target_sheet['{}{}:{}{}'.format(s_col, s_row, s_col, e_row)]:
            for cell in row:
                cell.border = Border(bottom=self.thin)
        for row in self.target_sheet['{}{}:{}{}'.format(s_col, s_row, e_col, s_row)]:
            for cell in row:
                cell.border = Border(bottom=self.thin)
        self.target_sheet['{}{}'.format(s_col, e_row)].border = Border(bottom=self.thin, left=self.thin)
        self.target_sheet['{}{}'.format(e_col, e_row)].border = Border(bottom=self.thin, right=self.thin)
        self.target_sheet['{}{}'.format(s_col, s_row)].border = Border(top=self.thin, left=self.thin)
        self.target_sheet['{}{}'.format(e_col, s_row)].border = Border(top=self.thin, right=self.thin)
