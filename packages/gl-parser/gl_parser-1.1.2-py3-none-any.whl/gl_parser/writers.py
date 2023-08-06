from pathlib import Path
from typing import Dict, Any, List, Tuple

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import Transaction
from .parsers import ParsingConfiguration


def write_page_headers(sheet: Worksheet, headers: List[str]):
    for i, header in enumerate(headers, 1):
        sheet.merge_cells(f'A{i}:L{i}')
        cell = sheet.cell(row=i, column=1)
        cell.value = header
        cell.alignment = Alignment(horizontal='center', vertical='center')


def write_parsed_transactions(target_file: Path, parsed_results: Dict[str, Any],
                              parsing_config: ParsingConfiguration):
    if target_file.exists():
        wb = load_workbook(target_file, data_only=True)
    else:
        wb = Workbook()
    for account_id in parsed_results['accounts'].keys():
        sheet = wb.create_sheet(account_id)
        # Write sheet headers
        write_page_headers(sheet, parsed_results['headers'])
        # Setting column widths
        for col_num, column_mapping in parsing_config.column_mappings.items():
            column_letter = get_column_letter(int(col_num))
            sheet.column_dimensions[column_letter].width = column_mapping.get('width', 20)
        # Write column headers
        write_column_headers(sheet, parsing_config.column_mappings, parsing_config.start_row)
        # Write transactions
        row = parsing_config.start_row
        transaction_list = parsed_results['accounts'][account_id]
        write_transactions(sheet, transaction_list, parsing_config.column_mappings, row)
    wb.save(target_file)


def write_transactions(sheet: Worksheet, transaction_list: List[Transaction],
                       column_mappings: Dict[str, Dict[str, Any]], row: int) -> int:
    for transaction in transaction_list:
        transaction_dict = transaction.dict()
        for col_num, col_mapping in column_mappings.items():
            sheet.cell(column=int(col_num), row=row, value=transaction_dict[col_mapping['name']])
            if col_mapping.get('number_format') is not None:
                sheet.cell(column=int(col_num), row=row).number_format = col_mapping.get('number_format')
        row += 1
    return row


def write_column_headers(sheet, column_mappings, start_row):
    for col_num, col_mapping in column_mappings.items():
        cell = sheet.cell(column=int(col_num), row=start_row - 1)
        cell.value = col_mapping['title']
        cell.font = Font(bold=True)


def write_fake_gl_file(excel_file: Path, parsed_results: Dict[str, Any],
                       parsing_config: ParsingConfiguration) -> Tuple[int, int]:
    wb = Workbook()
    sheet = wb.create_sheet(parsing_config.sheet_name)
    page_headers = parsed_results['headers']
    write_page_headers(sheet, page_headers)
    write_column_headers(sheet, parsing_config.column_mappings, parsing_config.start_row)
    row = parsing_config.start_row
    account_count = 0
    for _, transactions in parsed_results['accounts'].items():
        row = write_transactions(sheet, transactions, parsing_config.column_mappings, row)
        account_count += 1
    wb.save(excel_file)
    return account_count, row - parsing_config.start_row
