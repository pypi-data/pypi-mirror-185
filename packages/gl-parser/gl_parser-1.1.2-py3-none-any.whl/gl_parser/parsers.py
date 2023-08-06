import logging
from pathlib import Path
from typing import List, Any, Dict, Protocol

from openpyxl.reader.excel import load_workbook
from pydantic import ValidationError
from tqdm import tqdm

from gl_parser.exceptions import GLParserException
from gl_parser.models import Transaction

logger = logging.getLogger(__name__)


# COLUMN_MAPPINGS = {
#     1: {'name': 'account_id', 'title': 'Account ID', 'width': 12},
#     2: {'name': 'account_description', 'title': 'Account Description', 'width': 24},
#     3: {'name': 'date', 'title': 'Date', 'number_format': 'DD/MM/YYYY', 'width': 12},
#     4: {'name': 'reference', 'title': 'Reference', 'width': 30},
#     5: {'name': 'journal', 'title': 'Jrnl', 'width': 12},
#     6: {'name': 'description', 'title': 'Trans Description', 'width': 36},
#     7: {'name': 'debit_amount', 'title': 'Debit Amt', 'number_format': '#,##0.00', 'width': 12},
#     8: {'name': 'credit_amount', 'title': 'Credit Amt', 'number_format': '#,##0.00', 'width': 12},
#     9: {'name': 'balance', 'title': 'Balance', 'number_format': '#,##0.00', 'width': 12},
# }

# COLUMN_MAPPINGS_2 = {
#     '1': {'name': 'account_id', 'title': 'Account ID', 'width': 12},
#     '2': {'name': 'account_description', 'title': 'Account Description', 'width': 24},
#     '3': {'name': 'date', 'title': 'Date', 'number_format': 'DD/MM/YYYY', 'width': 12},
#     '4': {'name': 'reference', 'title': 'Reference', 'width': 30},
#     '5': {'name': 'journal', 'title': 'Jrnl', 'width': 12},
#     '6': {'name': 'description', 'title': 'Trans Description', 'width': 36},
#     '7': {'name': 'debit_amount', 'title': 'Debit Amt', 'number_format': '#,##0.00', 'width': 12},
#     '8': {'name': 'credit_amount', 'title': 'Credit Amt', 'number_format': '#,##0.00', 'width': 12},
#     '9': {'name': 'balance', 'title': 'Balance', 'number_format': '#,##0.00', 'width': 12},
# }


# def get_default_mappings():
#     return COLUMN_MAPPINGS


class ParsingConfiguration(Protocol):
    start_row: int
    sheet_name: str
    column_mappings: Dict[str, Dict[str, Any]]


def parse_general_ledger(general_ledger_file: Path,
                         parsing_config: ParsingConfiguration) -> Dict[str, Any]:
    accounts: Dict[str, List[Any]] = dict()
    headers = list()
    wb = load_workbook(general_ledger_file, data_only=True)
    try:
        sheet = wb[parsing_config.sheet_name]
    except KeyError:
        error_message = f'Sheet {parsing_config.sheet_name} not found in {general_ledger_file.name}'
        logger.error(error_message)
        raise GLParserException(error_message)
    # Read headers
    for row in range(1, parsing_config.start_row - 1):
        headers.append(sheet.cell(row=row, column=1).value)
    # Read account transactions
    last_row = sheet.max_row + 1
    current_account = None
    logger.debug(f'Parsing excel {general_ledger_file.name} from {parsing_config.start_row} to {last_row} row.')
    for row in tqdm(range(parsing_config.start_row, last_row)):
        transaction_dict = dict()
        account_id = sheet.cell(row=row, column=1).value
        if current_account is None:
            current_account = account_id
            accounts[account_id] = list()
        elif account_id == '':
            logger.debug(f'Blank account id in row {row} ')
            pass
        elif current_account != account_id and account_id is not None:
            # logger.debug(f'Row {row} current_account {current_account} != account_id {account_id}')
            current_account = account_id
            accounts[account_id] = list()
        transaction_obj = None
        for col, col_value in parsing_config.column_mappings.items():
            cell_obj = sheet.cell(row=row, column=int(col))
            value = cell_obj.value
            transaction_dict[col_value['name']] = value
        try:
            transaction_obj = Transaction(**transaction_dict)
        except ValidationError as e:
            error_message = f'Validation error on row {row}. Error: {e}'
            raise GLParserException(error_message)

        accounts[current_account].append(transaction_obj)
    parsed_results = {'headers': headers, 'accounts': accounts}
    return parsed_results
